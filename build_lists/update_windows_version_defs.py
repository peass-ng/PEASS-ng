#!/usr/bin/env python3

from __future__ import annotations

import argparse
import http.client
import json
import logging
import os
import re
import tempfile
import time
import zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen


BULLETIN_XLSX_URL = (
    "https://download.microsoft.com/download/6/7/3/"
    "673E4349-1CA5-40B9-8879-095C72D5B49D/BulletinSearch.xlsx"
)
MSRC_UPDATES_URL = "https://api.msrc.microsoft.com/cvrf/v3.0/updates"
MSRC_CVRF_ACCEPT = "application/json"
NVD_FEED_URL_TEMPLATE = "https://nvd.nist.gov/feeds/json/cve/2.0/nvdcve-2.0-{year}.json.zip"
USER_AGENT = "PEASS-ng windows_version_definitions updater"
KB_PATTERN = re.compile(r"\b(\d{6,7})\b")
WINDOWS_TOKEN = "windows"
LEGACY_PRODUCT_ALIASES: dict[str, tuple[tuple[str, ...], frozenset[str]]] = {
    "Microsoft Windows XP Service Pack 2": (
        ("Microsoft Windows XP", "Microsoft Windows XP Service Pack 1"),
        frozenset({"CVE-2017-0143"}),
    ),
    "Windows 10 for 32-bit Systems": (
        ("Windows 10 Version 1507 for 32-bit Systems",),
        frozenset({"CVE-2017-0143"}),
    ),
    "Windows 10 for x64-based Systems": (
        ("Windows 10 Version 1507 for x64-based Systems",),
        frozenset({"CVE-2017-0143"}),
    ),
    "Windows Server 2008 for 32-bit Systems Service Pack 2": (
        ("Windows Server 2008 for 32-bit Systems Service Pack 1",),
        frozenset({"CVE-2017-0143"}),
    ),
    "Windows Server 2008 for x64-based Systems Service Pack 2": (
        ("Windows Server 2008 for x64-based Systems Service Pack 1",),
        frozenset({"CVE-2017-0143"}),
    ),
}
LEGACY_COMPATIBILITY_ENTRIES: dict[str, tuple[dict[str, str], ...]] = {
    "Microsoft Windows XP": (
        {
            "cve": "CVE-2017-0143",
            "kb": "4012598",
            "severity": "Critical",
            "impact": "Remote Code Execution",
        },
    ),
    "Microsoft Windows XP Service Pack 1": (
        {
            "cve": "CVE-2017-0143",
            "kb": "4012598",
            "severity": "Critical",
            "impact": "Remote Code Execution",
        },
    ),
}


@dataclass(frozen=True)
class RawEntry:
    cve: str
    kb: str
    product: str
    severity: str
    impact: str
    supersedes: tuple[str, ...]


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate build_lists/windows_version_exploits.json directly from "
            "Microsoft Security Update Guide data, the legacy Microsoft bulletin "
            "workbook, and NVD exploit references."
        )
    )
    parser.add_argument(
        "--output",
        default=str(Path("build_lists") / "windows_version_exploits.json"),
    )
    parser.add_argument(
        "--msrc-max-workers",
        type=int,
        default=max(4, min(8, (os.cpu_count() or 4))),
        help="Maximum parallel downloads for MSRC CVRF documents.",
    )
    parser.add_argument(
        "--nvd-max-workers",
        type=int,
        default=max(4, min(8, (os.cpu_count() or 4))),
        help="Maximum parallel downloads for NVD yearly feeds.",
    )
    parser.add_argument(
        "--nvd-start-year",
        type=int,
        default=2002,
        help="First NVD year to process.",
    )
    parser.add_argument(
        "--nvd-end-year",
        type=int,
        default=datetime.now(timezone.utc).year,
        help="Last NVD year to process.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=180,
        help="Per-request timeout in seconds.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=4,
        help="Download retries for transient network failures.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable debug logging.",
    )
    return parser.parse_args()


def build_request(url: str, *, accept: str | None = None) -> Request:
    headers = {"User-Agent": USER_AGENT}
    if accept:
        headers["Accept"] = accept
    return Request(url, headers=headers)


def download_bytes(url: str, *, timeout: int, retries: int, accept: str | None = None) -> bytes:
    request = build_request(url, accept=accept)
    delay = 1.5
    for attempt in range(1, retries + 1):
        try:
            logging.debug("Downloading %s (attempt %d/%d)", url, attempt, retries)
            with urlopen(request, timeout=timeout) as response:
                payload = response.read()
                if not payload:
                    raise ValueError(f"Received an empty response from {url}")
                return payload
        except (HTTPError, URLError, TimeoutError, ValueError, http.client.IncompleteRead) as exc:
            if attempt == retries:
                raise RuntimeError(f"Failed to download {url}: {exc}") from exc
            logging.warning(
                "Download failed for %s on attempt %d/%d: %s",
                url,
                attempt,
                retries,
                exc,
            )
            time.sleep(delay)
            delay *= 2
    raise AssertionError("unreachable")


def download_json(url: str, *, timeout: int, retries: int, accept: str | None = None) -> Any:
    payload = download_bytes(url, timeout=timeout, retries=retries, accept=accept)
    try:
        return json.loads(payload.decode("utf-8"))
    except json.JSONDecodeError as exc:
        snippet = payload[:200].decode("utf-8", errors="replace")
        raise RuntimeError(
            f"Failed to decode JSON from {url}. Response starts with: {snippet!r}"
        ) from exc


def normalize_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def format_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")

    text = normalize_spaces(value)
    if not text:
        return ""

    for parser in (datetime.fromisoformat,):
        try:
            return parser(text.replace("Z", "+00:00")).strftime("%Y%m%d")
        except ValueError:
            pass

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            pass

    raise ValueError(f"Unsupported date value: {value!r}")


def extract_kbs(text: Any) -> list[str]:
    value = normalize_spaces(text)
    return KB_PATTERN.findall(value)


def get_latest_revision_date(vuln: dict[str, Any], fallback: str) -> str:
    latest = fallback
    for revision in vuln.get("RevisionHistory", []) or []:
        candidate = revision.get("Date")
        if not candidate:
            continue
        formatted = format_date(candidate)
        if formatted > latest:
            latest = formatted
    return latest


def find_note_title(notes: list[dict[str, Any]], target_type: str) -> str:
    for note in notes or []:
        note_type = str(note.get("Type", "")).strip()
        if note_type == target_type:
            title = note.get("Title")
            if isinstance(title, dict):
                return normalize_spaces(title.get("Value"))
            return normalize_spaces(title)
    return ""


def threat_type_matches(threat: dict[str, Any], target_type: str) -> bool:
    threat_type = threat.get("Type")
    if isinstance(threat_type, dict):
        threat_type = threat_type.get("Value")
    return str(threat_type).strip() == target_type


def get_threat_value(vuln: dict[str, Any], product_id: str, target_type: str) -> str:
    matches: list[str] = []
    fallback: list[str] = []

    for threat in vuln.get("Threats", []) or []:
        if not threat_type_matches(threat, target_type):
            continue
        description = threat.get("Description")
        if isinstance(description, dict):
            description = description.get("Value")
        value = normalize_spaces(description)
        if not value:
            continue
        product_ids = threat.get("ProductID") or []
        if isinstance(product_ids, str):
            product_ids = [product_ids]
        if product_id and product_id in product_ids:
            matches.append(value)
        else:
            fallback.append(value)

    if matches:
        return matches[0]
    if fallback:
        return fallback[0]
    return ""


def load_bulletin_entries(*, timeout: int, retries: int) -> list[RawEntry]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install it before running this generator."
        ) from exc

    logging.info("Downloading legacy Microsoft bulletin workbook")
    payload = download_bytes(BULLETIN_XLSX_URL, timeout=timeout, retries=retries)

    with tempfile.NamedTemporaryFile(prefix="bulletin_", suffix=".xlsx", delete=False) as handle:
        handle.write(payload)
        workbook_path = Path(handle.name)

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook.active
        entries: list[RawEntry] = []
        row_count = 0

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue

            row_count += 1
            if row_count % 5000 == 0:
                logging.info("Processed %d bulletin workbook rows", row_count)

            cves = [normalize_spaces(item) for item in str(row[13] or "").split(",") if normalize_spaces(item)]
            if not cves:
                continue

            date_posted = format_date(row[0])
            kb = normalize_spaces(row[7])
            product = normalize_spaces(row[6]).replace("2016 for x64-based Systems", "2016")
            severity = normalize_spaces(row[3])
            impact = normalize_spaces(row[4])
            supersedes = tuple(dict.fromkeys(extract_kbs(row[11])))

            for cve in cves:
                entries.append(
                    RawEntry(
                        cve=cve,
                        kb=kb,
                        product=product,
                        severity=severity,
                        impact=impact,
                        supersedes=supersedes,
                    )
                )

        workbook.close()
        logging.info("Collected %d raw bulletin entries", len(entries))
        return entries
    finally:
        workbook_path.unlink(missing_ok=True)


def fetch_msrc_update_catalog(*, timeout: int, retries: int) -> list[dict[str, Any]]:
    logging.info("Downloading Microsoft Security Update Guide update catalog")
    data = download_json(MSRC_UPDATES_URL, timeout=timeout, retries=retries)
    updates = data.get("value")
    if not isinstance(updates, list) or not updates:
        raise RuntimeError("MSRC updates catalog did not return a usable 'value' list")
    updates.sort(key=lambda item: item.get("InitialReleaseDate", ""))
    logging.info("Catalog contains %d MSRC update documents", len(updates))
    return updates


def fetch_msrc_document(url: str, *, timeout: int, retries: int) -> dict[str, Any]:
    return download_json(url, timeout=timeout, retries=retries, accept=MSRC_CVRF_ACCEPT)


def product_map_from_document(document: dict[str, Any]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for entry in document.get("ProductTree", {}).get("FullProductName", []) or []:
        product_id = normalize_spaces(entry.get("ProductID"))
        value = entry.get("Value")
        if isinstance(value, dict):
            value = value.get("Value")
        product_name = normalize_spaces(value)
        if product_id and product_name:
            mapping[product_id] = product_name
    return mapping


def extract_msrc_entries_from_document(document: dict[str, Any]) -> list[RawEntry]:
    entries: list[RawEntry] = []
    product_map = product_map_from_document(document)
    release_date = format_date(
        document.get("DocumentTracking", {}).get("InitialReleaseDate", datetime.now(timezone.utc))
    )

    for vuln in document.get("Vulnerability", []) or []:
        cve = normalize_spaces(vuln.get("CVE"))
        if not cve:
            continue

        posted = get_latest_revision_date(vuln, release_date)
        if not posted:
            posted = release_date

        for remediation in vuln.get("Remediations", []) or []:
            description = remediation.get("Description")
            if isinstance(description, dict):
                description = description.get("Value")
            description = normalize_spaces(description)
            kb_matches = extract_kbs(description)
            kb = kb_matches[0] if kb_matches else ""
            supersedes = tuple(dict.fromkeys(extract_kbs(remediation.get("Supercedence"))))
            product_ids = remediation.get("ProductID") or []
            if isinstance(product_ids, str):
                product_ids = [product_ids]

            for product_id in product_ids:
                product = product_map.get(normalize_spaces(product_id))
                if not product:
                    continue

                severity = get_threat_value(vuln, product_id, "3")
                impact = get_threat_value(vuln, product_id, "0")
                if not impact:
                    impact = find_note_title(vuln.get("Notes", []) or [], "7")

                entries.append(
                    RawEntry(
                        cve=cve,
                        kb=kb,
                        product=product,
                        severity=severity,
                        impact=impact,
                        supersedes=supersedes,
                    )
                )

    return entries


def load_msrc_entries(*, timeout: int, retries: int, max_workers: int) -> list[RawEntry]:
    updates = fetch_msrc_update_catalog(timeout=timeout, retries=retries)
    documents = [update["CvrfUrl"] for update in updates if normalize_spaces(update.get("CvrfUrl"))]
    entries: list[RawEntry] = []

    logging.info("Downloading %d MSRC CVRF documents with up to %d workers", len(documents), max_workers)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_url = {
            executor.submit(fetch_msrc_document, url, timeout=timeout, retries=retries): url
            for url in documents
        }
        completed = 0
        for future in as_completed(future_to_url):
            url = future_to_url[future]
            document = future.result()
            doc_entries = extract_msrc_entries_from_document(document)
            entries.extend(doc_entries)
            completed += 1
            if completed % 10 == 0 or completed == len(documents):
                logging.info(
                    "Processed %d/%d MSRC documents (%d cumulative entries)",
                    completed,
                    len(documents),
                    len(entries),
                )
            logging.debug("MSRC document %s produced %d raw entries", url, len(doc_entries))

    logging.info("Collected %d raw MSRC entries", len(entries))
    return entries


def extract_exploit_ids_from_feed(payload: bytes, *, year: int) -> set[str]:
    exploit_ids: set[str] = set()

    with tempfile.NamedTemporaryFile(prefix=f"nvdcve_{year}_", suffix=".zip", delete=False) as handle:
        handle.write(payload)
        archive_path = Path(handle.name)

    try:
        with zipfile.ZipFile(archive_path) as archive:
            json_name = next((name for name in archive.namelist() if name.endswith(".json")), None)
            if not json_name:
                raise RuntimeError(f"NVD archive for {year} does not contain a JSON file")
            with archive.open(json_name) as raw_json:
                document = json.load(raw_json)
    finally:
        archive_path.unlink(missing_ok=True)

    for item in document.get("vulnerabilities", []) or []:
        cve = item.get("cve", {})
        cve_id = normalize_spaces(cve.get("id"))
        if not cve_id:
            continue

        references = cve.get("references", []) or []
        for reference in references:
            tags = reference.get("tags") or []
            if "Exploit" in tags:
                exploit_ids.add(cve_id)
                break

    logging.debug("NVD %d contributed %d exploit-tagged CVEs", year, len(exploit_ids))
    return exploit_ids


def fetch_nvd_year(year: int, *, timeout: int, retries: int) -> set[str]:
    url = NVD_FEED_URL_TEMPLATE.format(year=year)
    logging.debug("Downloading NVD feed for %d", year)
    payload = download_bytes(url, timeout=timeout, retries=retries)
    return extract_exploit_ids_from_feed(payload, year=year)


def load_nvd_exploit_ids(
    *,
    start_year: int,
    end_year: int,
    timeout: int,
    retries: int,
    max_workers: int,
) -> set[str]:
    if start_year > end_year:
        raise ValueError(f"Invalid NVD year range: {start_year} > {end_year}")

    exploit_ids: set[str] = set()
    years = list(range(start_year, end_year + 1))
    logging.info("Downloading %d NVD yearly feeds with up to %d workers", len(years), max_workers)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_year = {
            executor.submit(fetch_nvd_year, year, timeout=timeout, retries=retries): year
            for year in years
        }
        completed = 0
        for future in as_completed(future_to_year):
            year = future_to_year[future]
            year_ids = future.result()
            exploit_ids.update(year_ids)
            completed += 1
            logging.info(
                "Processed NVD year %d (%d/%d, %d CVEs with exploit references, %d cumulative)",
                year,
                completed,
                len(years),
                len(year_ids),
                len(exploit_ids),
            )

    return exploit_ids


def build_definitions(entries: list[RawEntry], exploit_cves: set[str], generated: str) -> dict[str, Any]:
    products: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
    kb_supersedes: dict[str, set[str]] = defaultdict(set)

    for entry in entries:
        product = normalize_spaces(entry.product)
        if WINDOWS_TOKEN not in product.lower():
            continue

        kb = normalize_spaces(entry.kb)
        if kb:
            for superseded in entry.supersedes:
                if superseded:
                    kb_supersedes[kb].add(superseded)

        if not entry.cve or entry.cve not in exploit_cves:
            continue

        vuln_key = entry.cve or f"KB{kb}"
        if vuln_key in products[product]:
            continue

        products[product][vuln_key] = {
            "cve": entry.cve,
            "kb": kb,
            "severity": normalize_spaces(entry.severity),
            "impact": normalize_spaces(entry.impact),
        }

    for source_product, (aliases, allowed_keys) in LEGACY_PRODUCT_ALIASES.items():
        source_entries = products.get(source_product)
        if not source_entries:
            continue
        aliased_entries = {
            key: value
            for key, value in source_entries.items()
            if not allowed_keys or key in allowed_keys
        }
        if not aliased_entries:
            continue
        for alias in aliases:
            products.setdefault(alias, dict(aliased_entries))

    for product, entries_to_add in LEGACY_COMPATIBILITY_ENTRIES.items():
        destination = products.setdefault(product, {})
        for entry in entries_to_add:
            destination.setdefault(entry["cve"], dict(entry))

    data = {"generated": generated, "products": {}, "kb_supersedes": {}}
    for product in sorted(products):
        data["products"][product] = [products[product][key] for key in sorted(products[product])]
    for kb in sorted(kb_supersedes):
        data["kb_supersedes"][kb] = sorted(kb_supersedes[kb])
    return data


def validate_output(data: dict[str, Any]) -> None:
    if not re.fullmatch(r"\d{8}", str(data.get("generated", ""))):
        raise RuntimeError("Generated date is missing or malformed")
    products = data.get("products")
    if not isinstance(products, dict) or not products:
        raise RuntimeError("Output does not contain any product definitions")
    kb_supersedes = data.get("kb_supersedes")
    if not isinstance(kb_supersedes, dict):
        raise RuntimeError("Output does not contain a kb_supersedes mapping")

    sample_product = next(iter(products.values()))
    if not isinstance(sample_product, list):
        raise RuntimeError("Product entries must be lists")
    if sample_product:
        sample_entry = sample_product[0]
        required_keys = {"cve", "kb", "severity", "impact"}
        if not required_keys.issubset(sample_entry):
            raise RuntimeError(
                f"Product entries are missing keys. Required {sorted(required_keys)}, got {sorted(sample_entry)}"
            )


def main() -> None:
    args = parse_args()
    configure_logging(args.verbose)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logging.info("Starting windows version definition refresh")
    logging.info("Output path: %s", output_path)

    bulletin_entries = load_bulletin_entries(timeout=args.timeout, retries=args.retries)
    msrc_entries = load_msrc_entries(
        timeout=args.timeout,
        retries=args.retries,
        max_workers=args.msrc_max_workers,
    )
    exploit_cves = load_nvd_exploit_ids(
        start_year=args.nvd_start_year,
        end_year=args.nvd_end_year,
        timeout=args.timeout,
        retries=args.retries,
        max_workers=args.nvd_max_workers,
    )

    logging.info(
        "Building final JSON from %d bulletin entries, %d MSRC entries, and %d exploit-tagged CVEs",
        len(bulletin_entries),
        len(msrc_entries),
        len(exploit_cves),
    )

    generated = datetime.now(timezone.utc).strftime("%Y%m%d")
    data = build_definitions(bulletin_entries + msrc_entries, exploit_cves, generated)
    validate_output(data)

    output_path.write_text(json.dumps(data, separators=(",", ":")) + "\n", encoding="utf-8")

    total_products = len(data["products"])
    total_entries = sum(len(items) for items in data["products"].values())
    total_supersedes = len(data["kb_supersedes"])
    logging.info(
        "Generated %s (date=%s, products=%d, vulnerabilities=%d, supersedence_roots=%d)",
        output_path,
        data["generated"],
        total_products,
        total_entries,
        total_supersedes,
    )


if __name__ == "__main__":
    main()
